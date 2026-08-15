from collections import defaultdict
from datetime import timedelta
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import Count, Q, Sum, F, Avg, Value, DecimalField
from django.db.models.functions import Coalesce, TruncDate, TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.timezone import now
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from main import models
from main.validators import validate_image_file


def staff_required(view_func):
    """
    Faqat tizim ma'murlari (is_staff yoki is_superuser) uchun ruxsat beruvchi dekorator.
    """
    def check_user(u):
        return u.is_authenticated and (u.is_staff or u.is_superuser)
    return user_passes_test(check_user, login_url='d_login')(view_func)


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_admin_action(request, action, details=""):
    """
    Admin amallarini AuditLog jadvaliga yozadi.
    """
    try:
        models.AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=action,
            details=details,
            ip_address=get_client_ip(request)
        )
    except Exception:
        pass


# ==========================================
# 1. DASHBOARD BOSH SAHIFA (OVERVIEW)
# ==========================================

@staff_required
def index(request):
    current_time = now()
    today_start = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = current_time - timedelta(days=7)
    month_start = current_time - timedelta(days=30)

    # 1. Sales & Revenue KPIs (Delivered orders status=4)
    delivered_products = models.CartProduct.objects.filter(
        cart__status=4,
        product__isnull=False
    ).select_related('product', 'cart')

    total_income = sum(float(item.total_price) for item in delivered_products)

    today_income = sum(
        float(item.total_price) for item in delivered_products
        if item.cart and item.cart.date and item.cart.date >= today_start
    )

    week_income = sum(
        float(item.total_price) for item in delivered_products
        if item.cart and item.cart.date and item.cart.date >= week_start
    )

    month_income = sum(
        float(item.total_price) for item in delivered_products
        if item.cart and item.cart.date and item.cart.date >= month_start
    )

    completed_orders_count = models.Cart.objects.filter(status=4).count()
    aov = (total_income / completed_orders_count) if completed_orders_count > 0 else 0.0

    # 2. Orders KPIs
    total_orders_qs = models.Cart.objects.exclude(status=1)
    new_orders_count = models.Cart.objects.filter(status=2).count()
    processing_orders_count = models.Cart.objects.filter(status=3).count()
    completed_orders_count = models.Cart.objects.filter(status=4).count()
    cancelled_orders_count = models.Cart.objects.filter(status=5).count()

    # 3. Customers KPIs
    total_customers = models.User.objects.count()
    new_customers_month = models.User.objects.filter(date_joined__gte=month_start).count()
    active_customers = models.User.objects.filter(is_active=True).count()
    staff_users = models.User.objects.filter(is_staff=True).count()

    # 4. Inventory KPIs
    total_products = models.Product.objects.count()
    in_stock_products = models.Product.objects.filter(count__gt=5).count()
    low_stock_products = models.Product.objects.filter(count__gt=0, count__lte=5).count()
    out_of_stock_products = models.Product.objects.filter(count__lte=0).count()

    stats = {
        'total_income': total_income,
        'today_income': today_income,
        'week_income': week_income,
        'month_income': month_income,
        'aov': aov,
        'total_orders': total_orders_qs.count(),
        'new_orders': new_orders_count,
        'processing_orders': processing_orders_count,
        'completed_orders': completed_orders_count,
        'cancelled_orders': cancelled_orders_count,
        'total_customers': total_customers,
        'new_customers': new_customers_month,
        'active_customers': active_customers,
        'staff_users': staff_users,
        'total_products': total_products,
        'total_categories': models.Category.objects.count(),
        'in_stock_products': in_stock_products,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
    }

    recent_orders = (
        models.Cart.objects.exclude(status=1)
        .select_related('user')
        .prefetch_related('cartproduct_set__product')
        .order_by('-date', '-id')[:8]
    )

    # Top selling products in delivered orders
    top_selling_items = (
        models.CartProduct.objects.filter(cart__status=4, product__isnull=False)
        .values('product_id', 'product__name', 'product__image', 'product__price', 'product__category__name')
        .annotate(total_sold=Sum('count'))
        .order_by('-total_sold')[:5]
    )

    # Low stock alerts list
    low_stock_list = models.Product.objects.filter(count__lte=5).select_related('category').order_by('count')[:6]

    recent_users = models.User.objects.order_by('-date_joined')[:6]

    return render(request, 'dashboard/index.html', {
        'stats': stats,
        'recent_orders': recent_orders,
        'top_selling_items': top_selling_items,
        'low_stock_list': low_stock_list,
        'recent_users': recent_users,
    })


# ==========================================
# 2. SAVDO TAHLILI (SALES ANALYTICS)
# ==========================================

@staff_required
def sales_analytics(request):
    period = request.GET.get('period', '30days')
    current_time = now()

    if period == 'today':
        start_date = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = "Bugungi kun"
    elif period == '7days':
        start_date = current_time - timedelta(days=7)
        period_label = "So'nggi 7 kun"
    elif period == '90days':
        start_date = current_time - timedelta(days=90)
        period_label = "So'nggi 90 kun"
    elif period == '12months':
        start_date = current_time - timedelta(days=365)
        period_label = "So'nggi 12 oy"
    elif period == 'all':
        start_date = None
        period_label = "Barcha davr"
    else:  # default 30days
        period = '30days'
        start_date = current_time - timedelta(days=30)
        period_label = "So'nggi 30 kun"

    delivered_items = models.CartProduct.objects.filter(
        cart__status=4,
        product__isnull=False
    ).select_related('product', 'product__category', 'cart')

    if start_date:
        delivered_items = delivered_items.filter(cart__date__gte=start_date)

    # Calculate Period Total Revenue and Items Sold
    period_revenue = sum(float(item.total_price) for item in delivered_items)
    period_items_sold = sum(item.count for item in delivered_items)

    period_orders_qs = models.Cart.objects.filter(status=4)
    if start_date:
        period_orders_qs = period_orders_qs.filter(date__gte=start_date)
    period_orders_count = period_orders_qs.count()
    period_aov = (period_revenue / period_orders_count) if period_orders_count > 0 else 0.0

    # Sales by Category
    category_revenue_map = defaultdict(lambda: {'name': '', 'revenue': 0.0, 'count': 0})
    for item in delivered_items:
        cat_name = item.product.category.name if item.product and item.product.category else "Boshqa"
        category_revenue_map[cat_name]['name'] = cat_name
        category_revenue_map[cat_name]['revenue'] += float(item.total_price)
        category_revenue_map[cat_name]['count'] += item.count

    category_sales = sorted(category_revenue_map.values(), key=lambda x: x['revenue'], reverse=True)
    for cat in category_sales:
        cat['percentage'] = round((cat['revenue'] / period_revenue * 100), 1) if period_revenue > 0 else 0.0

    # Top 10 Products by revenue in period
    product_revenue_map = defaultdict(lambda: {'name': '', 'image': '', 'category': '', 'price': 0, 'units': 0, 'revenue': 0.0})
    for item in delivered_items:
        pid = item.product.id
        product_revenue_map[pid]['name'] = item.product.name
        product_revenue_map[pid]['image'] = item.product.image.url if item.product.image else ''
        product_revenue_map[pid]['category'] = item.product.category.name if item.product.category else ''
        product_revenue_map[pid]['price'] = float(item.product.active_price)
        product_revenue_map[pid]['units'] += item.count
        product_revenue_map[pid]['revenue'] += float(item.total_price)

    top_products = sorted(product_revenue_map.values(), key=lambda x: x['revenue'], reverse=True)[:10]

    context = {
        'period': period,
        'period_label': period_label,
        'period_revenue': period_revenue,
        'period_orders_count': period_orders_count,
        'period_items_sold': period_items_sold,
        'period_aov': period_aov,
        'category_sales': category_sales,
        'top_products': top_products,
    }
    return render(request, 'dashboard/analytics.html', context)


# ==========================================
# 3. KATEGORIYA BOSHQARUVI
# ==========================================

@staff_required
def create_category(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            logo = request.FILES.get('logo')
            is_active = bool(request.POST.get('is_active', True))

            if not name or not logo:
                messages.warning(request, "Kategoriya nomi va logotipini kiritish majburiy.")
                return render(request, 'dashboard/create_category.html')

            validate_image_file(logo)
            cat = models.Category.objects.create(name=name, logo=logo, is_active=is_active)
            log_admin_action(request, "CATEGORY_CREATE", f"Kategoriya yaratildi: {cat.name} (ID: {cat.id})")
            messages.success(request, f"'{name}' kategoriyasi muvaffaqiyatli yaratildi.")
            return redirect('d_list_category')
        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {e}")
            return render(request, 'dashboard/create_category.html')

    return render(request, 'dashboard/create_category.html')


@staff_required
def list_category(request):
    query = request.GET.get('query', '').strip()
    status_filter = request.GET.get('status')

    categories = models.Category.objects.annotate(
        product_count=Count('product')
    ).order_by('-id')

    if query:
        categories = categories.filter(name__icontains=query)

    if status_filter == 'active':
        categories = categories.filter(is_active=True)
    elif status_filter == 'inactive':
        categories = categories.filter(is_active=False)

    paginator = Paginator(categories, 12)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    return render(request, 'dashboard/category_list.html', {
        'categories': page_obj,
        'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
    })


@staff_required
def edit_category(request, id):
    category = get_object_or_404(models.Category, id=id)
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            if not name:
                messages.warning(request, "Kategoriya nomi bo'sh bo'lishi mumkin emas.")
                return render(request, 'dashboard/edit_category.html', {'category': category})

            category.name = name
            category.is_active = bool(request.POST.get('is_active'))

            logo = request.FILES.get('logo')
            if logo:
                validate_image_file(logo)
                category.logo = logo

            category.save()
            log_admin_action(request, "CATEGORY_UPDATE", f"Kategoriya yangilandi: {category.name} (ID: {category.id})")
            messages.success(request, f"'{category.name}' kategoriyasi muvaffaqiyatli yangilandi.")
            return redirect('d_list_category')
        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {e}")

    return render(request, 'dashboard/edit_category.html', {'category': category})


@staff_required
def delete_category(request, id):
    category = get_object_or_404(models.Category, id=id)
    product_count = category.product_set.count()
    category_name = category.name
    category.delete()
    log_admin_action(request, "CATEGORY_DELETE", f"Kategoriya o'chirildi: {category_name} ({product_count} ta mahsulot)")
    if product_count > 0:
        messages.warning(request, f"'{category_name}' kategoriyasi va unga tegishli {product_count} ta mahsulot o'chirildi.")
    else:
        messages.success(request, f"'{category_name}' kategoriyasi o'chirildi.")
    return redirect('d_list_category')


# ==========================================
# 4. MAHSULOT VA OMBOR BOSHQARUVI (PRODUCTS & INVENTORY)
# ==========================================

@staff_required
def create_product(request):
    categories = models.Category.objects.all().order_by('name')
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            category_id = request.POST.get('category')
            description = request.POST.get('description', '').strip()
            price = request.POST.get('price')
            discount_price = request.POST.get('discount_price')
            image = request.FILES.get('image')
            discount_status = bool(request.POST.get('discount_status'))
            count = request.POST.get('count', 0)

            if not name or not category_id or not description or not price or not image:
                messages.warning(request, "Barcha majburiy maydonlarni to'ldiring va rasm yuklang.")
                return render(request, 'dashboard/create_praduct.html', {'categories': categories})

            validate_image_file(image)
            category = get_object_or_404(models.Category, id=category_id)

            price_val = float(price)
            disc_price_val = float(discount_price) if discount_price else None

            if disc_price_val is not None and disc_price_val >= price_val:
                messages.warning(request, "Chegirma narxi asosiy narxdan kichik bo'lishi kerak.")
                return render(request, 'dashboard/create_praduct.html', {'categories': categories})

            prod = models.Product.objects.create(
                name=name,
                category=category,
                description=description,
                price=price_val,
                discount_price=disc_price_val,
                image=image,
                discount_status=discount_status,
                count=int(count) if count else 0
            )
            log_admin_action(request, "PRODUCT_CREATE", f"Mahsulot yaratildi: {prod.name} (Code: {prod.code}, Narx: {prod.price})")
            messages.success(request, f"'{name}' mahsuloti muvaffaqiyatli yaratildi.")
            return redirect('d_list_product')
        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")
            return render(request, 'dashboard/create_praduct.html', {'categories': categories})

    return render(request, 'dashboard/create_praduct.html', {'categories': categories})


@staff_required
def list_product(request):
    query = request.GET.get('query', '').strip()
    category_id = request.GET.get('category_id')
    discount_filter = request.GET.get('discount')
    stock_filter = request.GET.get('stock')
    ordering = request.GET.get('ordering', '-created_at')

    products = models.Product.objects.select_related('category').all()

    if query:
        products = products.filter(
            Q(name__icontains=query) | Q(code__icontains=query) | Q(description__icontains=query)
        )

    if category_id:
        products = products.filter(category_id=category_id)

    if discount_filter == '1':
        products = products.filter(discount_status=True)

    if stock_filter == 'in_stock':
        products = products.filter(count__gt=5)
    elif stock_filter == 'low_stock':
        products = products.filter(count__gt=0, count__lte=5)
    elif stock_filter == 'out_of_stock':
        products = products.filter(count__lte=0)

    # Sorting
    valid_orderings = ['-created_at', 'created_at', 'price', '-price', 'count', '-count', 'name']
    if ordering in valid_orderings:
        products = products.order_by(ordering)
    else:
        products = products.order_by('-created_at')

    paginator = Paginator(products, 15)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    categories = models.Category.objects.all().order_by('name')

    return render(request, 'dashboard/praduct_list.html', {
        'products': page_obj,
        'page_obj': page_obj,
        'query': query,
        'categories': categories,
        'selected_category': category_id,
        'selected_discount': discount_filter,
        'stock_filter': stock_filter,
        'ordering': ordering,
    })


@staff_required
def edit_product(request, code):
    product = get_object_or_404(models.Product, code=code)
    categories = models.Category.objects.all().order_by('name')

    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            category_id = request.POST.get('category')
            description = request.POST.get('description', '').strip()
            price = request.POST.get('price')
            discount_price = request.POST.get('discount_price')
            discount_status = bool(request.POST.get('discount_status'))
            count = request.POST.get('count', 0)

            if not name or not price:
                messages.warning(request, "Mahsulot nomi va narxi bo'sh bo'lishi mumkin emas.")
                return render(request, 'dashboard/edit_praduct.html', {'product': product, 'categories': categories})

            product.name = name
            if category_id:
                product.category = get_object_or_404(models.Category, id=category_id)

            product.description = description
            price_val = float(price)
            disc_price_val = float(discount_price) if discount_price else None

            if disc_price_val is not None and disc_price_val >= price_val:
                messages.warning(request, "Chegirma narxi asosiy narxdan kichik bo'lishi kerak.")
                return render(request, 'dashboard/edit_praduct.html', {'product': product, 'categories': categories})

            product.price = price_val
            product.discount_price = disc_price_val
            product.discount_status = discount_status
            product.count = int(count) if count else 0

            image = request.FILES.get('image')
            if image:
                validate_image_file(image)
                product.image = image

            product.save()
            log_admin_action(request, "PRODUCT_UPDATE", f"Mahsulot yangilandi: {product.name} (Code: {product.code})")
            messages.success(request, f"'{product.name}' mahsuloti muvaffaqiyatli yangilandi.")
            return redirect('d_list_product')
        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")

    context = {'product': product, 'categories': categories}
    return render(request, 'dashboard/edit_praduct.html', context=context)


@staff_required
def delete_product(request, code):
    product = get_object_or_404(models.Product, code=code)
    prod_name = product.name
    product.delete()
    log_admin_action(request, "PRODUCT_DELETE", f"Mahsulot o'chirildi: {prod_name} (Code: {code})")
    messages.success(request, f"'{prod_name}' mahsuloti o'chirildi.")
    return redirect('d_list_product')


@staff_required
def inventory_list(request):
    """
    Ombor qoldiqlari nazorati (Inventory Management).
    """
    query = request.GET.get('query', '').strip()
    status_filter = request.GET.get('status')
    category_id = request.GET.get('category_id')

    products = models.Product.objects.select_related('category').all().order_by('count', 'name')

    if query:
        products = products.filter(Q(name__icontains=query) | Q(code__icontains=query))

    if category_id:
        products = products.filter(category_id=category_id)

    if status_filter == 'in_stock':
        products = products.filter(count__gt=5)
    elif status_filter == 'low_stock':
        products = products.filter(count__gt=0, count__lte=5)
    elif status_filter == 'out_of_stock':
        products = products.filter(count__lte=0)

    total_products_count = models.Product.objects.count()
    total_stock_units = models.Product.objects.aggregate(total=Sum('count'))['total'] or 0
    in_stock_count = models.Product.objects.filter(count__gt=5).count()
    low_stock_count = models.Product.objects.filter(count__gt=0, count__lte=5).count()
    out_of_stock_count = models.Product.objects.filter(count__lte=0).count()

    paginator = Paginator(products, 20)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    categories = models.Category.objects.all().order_by('name')

    return render(request, 'dashboard/inventory.html', {
        'products': page_obj,
        'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
        'category_id': category_id,
        'categories': categories,
        'total_products_count': total_products_count,
        'total_stock_units': total_stock_units,
        'in_stock_count': in_stock_count,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
    })


@staff_required
@require_POST
def update_stock(request, code):
    """
    Ombordagi qoldiqni tezkor o'zgartirish.
    """
    product = get_object_or_404(models.Product, code=code)
    try:
        new_count = int(request.POST.get('count', 0))
        if new_count < 0:
            messages.error(request, "Ombor qoldig'i manfiy bo'lishi mumkin emas.")
            return redirect('d_inventory')

        old_count = product.count
        product.count = new_count
        product.save(update_fields=['count'])

        log_admin_action(request, "STOCK_UPDATE", f"Mahsulot ombor qoldig'i o'zgartirildi: {product.name} ({old_count} -> {new_count})")
        messages.success(request, f"'{product.name}' ombor qoldig'i yangilandi: {new_count} dona.")
    except ValueError:
        messages.error(request, "Noto'g'ri son kiritildi.")
    return redirect('d_inventory')


# ==========================================
# 5. BUYURTMALAR BOSHQARUVI (ORDER MANAGEMENT)
# ==========================================

@staff_required
def orders(request):
    query = request.GET.get('query', '').strip()
    status_filter = request.GET.get('status')
    period = request.GET.get('period', 'all')
    ordering = request.GET.get('ordering', '-date')

    order_qs = (
        models.Cart.objects.exclude(status=1)
        .select_related('user')
        .prefetch_related('cartproduct_set__product')
    )

    if query:
        order_qs = order_qs.filter(
            Q(code__icontains=query) |
            Q(user__username__icontains=query) |
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__phone__icontains=query)
        )

    if status_filter and status_filter.isdigit():
        order_qs = order_qs.filter(status=int(status_filter))

    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    if period == 'today':
        order_qs = order_qs.filter(date__gte=today_start)
    elif period == 'yesterday':
        yesterday_start = today_start - timedelta(days=1)
        order_qs = order_qs.filter(date__gte=yesterday_start, date__lt=today_start)
    elif period == '7days':
        order_qs = order_qs.filter(date__gte=now - timedelta(days=7))
    elif period == '30days':
        order_qs = order_qs.filter(date__gte=now - timedelta(days=30))
    elif period == 'this_month':
        order_qs = order_qs.filter(date__year=now.year, date__month=now.month)
    elif period == 'last_month':
        first_day_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        order_qs = order_qs.filter(date__year=last_day_last_month.year, date__month=last_day_last_month.month)
    elif period == 'this_year':
        order_qs = order_qs.filter(date__year=now.year)
    elif period and period.startswith('month_'):
        try:
            parts = period.replace('month_', '').split('-')
            if len(parts) == 2:
                y, m = int(parts[0]), int(parts[1])
                order_qs = order_qs.filter(date__year=y, date__month=m)
        except Exception:
            pass

    if ordering == 'date_asc':
        order_qs = order_qs.order_by('date', 'id')
    else:
        order_qs = order_qs.order_by('-date', '-id')

    # Available sales months from DB
    month_names_uz = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
        5: "May", 6: "Iyun", 7: "Iyul", 8: "Avgust",
        9: "Sentyabr", 10: "Oktyabr", 11: "Noyabr", 12: "Dekabr"
    }

    try:
        available_months_raw = (
            models.Cart.objects.exclude(status=1)
            .dates('date', 'month', order='DESC')
        )
        available_months = []
        for m_date in available_months_raw:
            val = f"month_{m_date.year}-{m_date.month:02d}"
            label = f"{month_names_uz.get(m_date.month, m_date.strftime('%B'))} {m_date.year}"
            available_months.append({
                'value': val,
                'label': label
            })
    except Exception:
        available_months = []

    # Status summary counts
    status_counts = {
        'all': models.Cart.objects.exclude(status=1).count(),
        'new': models.Cart.objects.filter(status=2).count(),
        'processing': models.Cart.objects.filter(status=3).count(),
        'completed': models.Cart.objects.filter(status=4).count(),
        'cancelled': models.Cart.objects.filter(status=5).count(),
    }

    paginator = Paginator(order_qs, 15)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    return render(request, 'dashboard/order_list.html', {
        'orders': page_obj,
        'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
        'period': period,
        'ordering': ordering,
        'available_months': available_months,
        'status_counts': status_counts,
    })


@staff_required
def status_update(request, code):
    order = get_object_or_404(models.Cart, code=code)
    target_status = request.POST.get('target_status') if request.method == 'POST' else request.GET.get('target_status')
    comment = request.POST.get('comment', '').strip() if request.method == 'POST' else ''

    status_labels = {
        2: "Qabul qilindi (Yangi)",
        3: "Yo'lda / Jarayonda",
        4: "Yetkazilgan",
        5: "Bekor qilingan / Qaytarilgan"
    }

    if target_status and target_status.isdigit():
        new_st = int(target_status)
        if new_st in (2, 3, 4, 5) and new_st != order.status:
            old_st = order.status
            with transaction.atomic():
                # If cancelling order (status 5), replenish stock
                if new_st == 5 and old_st in (2, 3, 4):
                    for cp in order.cartproduct_set.filter(product__isnull=False):
                        models.Product.objects.filter(pk=cp.product.pk).update(count=F('count') + cp.count)

                # If re-activating a previously cancelled order, deduct stock
                elif old_st == 5 and new_st in (2, 3, 4):
                    for cp in order.cartproduct_set.filter(product__isnull=False):
                        models.Product.objects.filter(pk=cp.product.pk).update(count=F('count') - cp.count)

                order.status = new_st
                order.save(update_fields=['status'])

                models.OrderStatusHistory.objects.create(
                    order=order,
                    old_status=old_st,
                    new_status=new_st,
                    changed_by=request.user,
                    comment=comment or f"Status {status_labels.get(new_st, new_st)} ga o'zgartirildi."
                )

                log_admin_action(
                    request,
                    "ORDER_STATUS_UPDATE",
                    f"Buyurtma #{str(order.code)[:8]} statusi: {old_st} -> {new_st}. Izoh: {comment}"
                )

            messages.success(request, f"Buyurtma #{str(order.code)[:8]} statusi '{status_labels.get(new_st)}' ga o'zgartirildi.")
            return redirect(request.META.get('HTTP_REFERER', 'd_orders'))

    # Fallback to next step
    if order.status in (2, 3):
        old_st = order.status
        new_st = order.status + 1
        with transaction.atomic():
            order.status = new_st
            order.save(update_fields=['status'])
            models.OrderStatusHistory.objects.create(
                order=order,
                old_status=old_st,
                new_status=new_st,
                changed_by=request.user,
                comment=f"Keyingi bosqich: {status_labels.get(new_st)}"
            )
            log_admin_action(request, "ORDER_STATUS_UPDATE", f"Buyurtma #{str(order.code)[:8]}: {old_st} -> {new_st}")
        messages.success(request, f"Buyurtma #{str(order.code)[:8]} keyingi bosqichga o'tkazildi.")
    else:
        messages.warning(request, "Ushbu buyurtma statusini avtomatik oshirib bo'lmaydi.")

    return redirect(request.META.get('HTTP_REFERER', 'd_orders'))


@staff_required
def reject_cart(request, code):
    order = get_object_or_404(models.Cart, code=code)
    if order.status in (2, 3, 4):
        old_st = order.status
        with transaction.atomic():
            # Replenish stock
            for cp in order.cartproduct_set.filter(product__isnull=False):
                models.Product.objects.filter(pk=cp.product.pk).update(count=F('count') + cp.count)

            order.status = 5
            order.save(update_fields=['status'])

            models.OrderStatusHistory.objects.create(
                order=order,
                old_status=old_st,
                new_status=5,
                changed_by=request.user,
                comment="Admin tomonidan bekor qilindi / qaytarildi."
            )
            log_admin_action(request, "ORDER_REJECT", f"Buyurtma #{str(order.code)[:8]} bekor qilindi (Qaytarildi)")

        messages.success(request, f"Buyurtma #{str(order.code)[:8]} bekor qilindi va mahsulotlar omborga qaytarildi.")
    else:
        messages.warning(request, "Ushbu buyurtmani bekor qilib bo'lmaydi.")
    return redirect(request.META.get('HTTP_REFERER', 'd_orders'))


@staff_required
def cart_detail(request, code):
    order = get_object_or_404(
        models.Cart.objects.select_related('user'),
        code=code
    )
    cart_products = models.CartProduct.objects.filter(cart=order).select_related('product', 'product__category')

    total_amount = sum(cp.total_price for cp in cart_products)
    total_count = sum(cp.count for cp in cart_products)

    from main.services.payment import PaymentManager
    financials = PaymentManager.calculate_order_financials(order)

    # Customer Lifetime Statistics
    customer_orders_count = models.Cart.objects.filter(user=order.user).exclude(status=1).count()
    customer_completed_products = models.CartProduct.objects.filter(
        cart__user=order.user, cart__status=4, product__isnull=False
    ).select_related('product')
    customer_lifetime_spent = sum(cp.total_price for cp in customer_completed_products)

    # Status History
    status_history = order.status_history.select_related('changed_by').all()
    payments = order.payments.all().order_by('-created_at')
    primary_payment = payments.first()

    context = {
        'order': order,
        'cart_products': cart_products,
        'total_amount': total_amount,
        'total_count': total_count,
        'financials': financials,
        'customer_orders_count': customer_orders_count,
        'customer_lifetime_spent': customer_lifetime_spent,
        'status_history': status_history,
        'payments': payments,
        'primary_payment': primary_payment,
    }
    return render(request, 'dashboard/orders_detail.html', context=context)


@staff_required
@require_POST
def settle_order_balance(request, code):
    """
    Buyurtmaning qoldiq summasini yetkazib berishda (naqd/terminal) qabul qilib to'liq to'langan deb belgilash.
    """
    order = get_object_or_404(models.Cart, code=code, status__gt=1)
    amount_raw = request.POST.get('amount')
    comment = request.POST.get('comment', '').strip()

    try:
        amount = Decimal(str(amount_raw)) if amount_raw else None
        from main.services.payment import PaymentManager
        PaymentManager.settle_cash_balance(
            order=order,
            amount=amount,
            user=request.user,
            comment=comment or "Yetkazib berishda naqd to'lov orqali qoldiq yopildi."
        )
        messages.success(request, f"Buyurtma #{str(order.code)[:8]} uchun qoldiq to'lov muvaffaqiyatli qabul qilindi!")
    except Exception as e:
        messages.error(request, f"Qoldiq to'lovni rasmiylashtirishda xatolik: {str(e)}")

    return redirect('d_detail_orders', code=order.code)


@staff_required
def payments_list(request):
    """
    To'lovlar monitoringi va boshqaruvi (Prepayments, Balance settlements, Refunds).
    """
    query = request.GET.get('query', '').strip()
    status_filter = request.GET.get('status', '').strip()
    provider_filter = request.GET.get('provider', '').strip()
    purpose_filter = request.GET.get('purpose', '').strip()

    payments_qs = models.Payment.objects.select_related('order', 'order__user').order_by('-created_at')

    if query:
        payments_qs = payments_qs.filter(
            Q(code__icontains=query) |
            Q(transaction_id__icontains=query) |
            Q(order__code__icontains=query) |
            Q(order__user__username__icontains=query) |
            Q(order__user__phone__icontains=query)
        )

    if status_filter:
        payments_qs = payments_qs.filter(status=status_filter)

    if provider_filter:
        payments_qs = payments_qs.filter(provider=provider_filter)

    if purpose_filter:
        payments_qs = payments_qs.filter(purpose=purpose_filter)

    # KPI Statistics
    total_payments_count = models.Payment.objects.count()
    paid_payments = models.Payment.objects.filter(status=models.Payment.Status.PAID)
    total_paid_amount = sum(float(p.amount) for p in paid_payments)
    prepayment_collected = sum(float(p.amount) for p in paid_payments.filter(purpose=models.Payment.Purpose.PREPAYMENT))
    balance_collected = sum(float(p.amount) for p in paid_payments.filter(purpose=models.Payment.Purpose.BALANCE))

    outstanding_orders = models.Cart.objects.filter(status__in=[2, 3])
    outstanding_balances = sum(float(o.remaining_amount) for o in outstanding_orders)

    paid_count = paid_payments.count()
    pending_count = models.Payment.objects.filter(status=models.Payment.Status.PENDING).count()
    failed_refunded_count = models.Payment.objects.filter(
        status__in=[models.Payment.Status.FAILED, models.Payment.Status.REFUNDED, models.Payment.Status.CANCELLED]
    ).count()

    paginator = Paginator(payments_qs, 20)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    context = {
        'payments': page_obj,
        'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
        'provider_filter': provider_filter,
        'purpose_filter': purpose_filter,
        'total_payments_count': total_payments_count,
        'total_paid_amount': total_paid_amount,
        'prepayment_collected': prepayment_collected,
        'balance_collected': balance_collected,
        'outstanding_balances': outstanding_balances,
        'paid_count': paid_count,
        'pending_count': pending_count,
        'failed_refunded_count': failed_refunded_count,
        'provider_choices': models.Payment.Provider.choices,
        'status_choices': models.Payment.Status.choices,
        'purpose_choices': models.Payment.Purpose.choices,
    }
    return render(request, 'dashboard/payments.html', context)


@staff_required
@require_POST
def refund_payment(request, payment_id):
    """
    To'lovni qaytarish (Admin Refund).
    """
    payment = get_object_or_404(models.Payment, id=payment_id)
    reason = request.POST.get('reason', '').strip()
    amount_raw = request.POST.get('amount')
    amount = float(amount_raw) if amount_raw else None

    from main.services.payment import PaymentManager
    result = PaymentManager.refund_payment(payment, amount=amount, reason=reason)
    if result.get('success'):
        log_admin_action(
            request,
            "PAYMENT_REFUND",
            f"To'lov #{str(payment.code)[:8]} qaytarildi. Summa: {payment.amount} {payment.currency}. Sabab: {reason}"
        )
        messages.success(request, f"To'lov #{str(payment.code)[:8]} muvaffaqiyatli qaytarildi.")
    else:
        messages.error(request, result.get('message', "To'lovni qaytarishda xatolik."))

    return redirect(request.META.get('HTTP_REFERER', 'd_payments'))


# ==========================================
# 6. FOYDALANUVCHILAR BOSHQARUVI (CUSTOMER MANAGEMENT)
# ==========================================


@staff_required
def list_users(request):
    query = request.GET.get('query', '').strip()
    role_filter = request.GET.get('role')

    users_qs = models.User.objects.annotate(
        order_count=Count('cart', filter=~Q(cart__status=1))
    ).order_by('-date_joined')

    if query:
        users_qs = users_qs.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(phone__icontains=query) |
            Q(email__icontains=query)
        )

    if role_filter == 'staff':
        users_qs = users_qs.filter(is_staff=True)
    elif role_filter == 'customer':
        users_qs = users_qs.filter(is_staff=False)

    paginator = Paginator(users_qs, 15)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    return render(request, 'dashboard/user_list.html', {
        'users': page_obj,
        'page_obj': page_obj,
        'query': query,
        'role_filter': role_filter,
    })


@staff_required
def customer_detail(request, id):
    customer = get_object_or_404(models.User, id=id)

    # Customer orders
    orders_list = (
        models.Cart.objects.filter(user=customer)
        .exclude(status=1)
        .prefetch_related('cartproduct_set__product')
        .order_by('-date')
    )

    delivered_items = models.CartProduct.objects.filter(
        cart__user=customer,
        cart__status=4,
        product__isnull=False
    ).select_related('product')

    total_spent = sum(cp.total_price for cp in delivered_items)
    completed_orders = orders_list.filter(status=4).count()
    aov = (total_spent / completed_orders) if completed_orders > 0 else 0.0

    context = {
        'customer': customer,
        'orders': orders_list,
        'total_orders': orders_list.count(),
        'completed_orders': completed_orders,
        'total_spent': total_spent,
        'aov': aov,
    }
    return render(request, 'dashboard/customer_detail.html', context)


@staff_required
def toggle_user_status(request, id):
    target_user = get_object_or_404(models.User, id=id)

    if target_user == request.user:
        messages.error(request, "O'z hisobingiz holatini o'zgartira olmaysiz.")
        return redirect('d_list_users')

    if target_user.is_superuser and not request.user.is_superuser:
        messages.error(request, "Superadmin holatini faqat boshqa superadmin o'zgartira oladi.")
        return redirect('d_list_users')

    target_user.is_active = not target_user.is_active
    target_user.save(update_fields=['is_active'])

    holat = "faollashtirildi" if target_user.is_active else "faolsizlantirildi"
    log_admin_action(request, "USER_STATUS_TOGGLE", f"Foydalanuvchi '{target_user.username}' {holat}")
    messages.success(request, f"Foydalanuvchi '{target_user.username}' muvaffaqiyatli {holat}.")
    return redirect(request.META.get('HTTP_REFERER', 'd_list_users'))


# ==========================================
# 7. HISOBOTLAR MODULI (REPORTS & EXCEL EXPORTS)
# ==========================================

@staff_required
def reports_overview(request):
    """
    Professional hisobotlar boshqaruv paneli.
    """
    # 1. Monthly sales summary
    delivered_products = models.CartProduct.objects.filter(
        cart__status=4,
        product__isnull=False
    ).select_related('product', 'cart')

    monthly_summary = defaultdict(lambda: {'orders': set(), 'items': 0, 'revenue': 0.0})
    for cp in delivered_products:
        if cp.cart and cp.cart.date:
            m_key = cp.cart.date.strftime('%Y-%m')
            monthly_summary[m_key]['orders'].add(cp.cart.id)
            monthly_summary[m_key]['items'] += cp.count
            monthly_summary[m_key]['revenue'] += float(cp.total_price)

    monthly_report = []
    for m_key in sorted(monthly_summary.keys(), reverse=True):
        monthly_report.append({
            'month': m_key,
            'orders_count': len(monthly_summary[m_key]['orders']),
            'items_count': monthly_summary[m_key]['items'],
            'revenue': monthly_summary[m_key]['revenue'],
        })

    # 2. Category Performance
    categories_perf = []
    for cat in models.Category.objects.all():
        cat_items = delivered_products.filter(product__category=cat)
        cat_rev = sum(float(cp.total_price) for cp in cat_items)
        cat_sold = sum(cp.count for cp in cat_items)
        categories_perf.append({
            'name': cat.name,
            'products_count': cat.product_set.count(),
            'units_sold': cat_sold,
            'revenue': cat_rev,
        })
    categories_perf.sort(key=lambda x: x['revenue'], reverse=True)

    # 3. Top Products Performance
    top_products_perf = (
        models.CartProduct.objects.filter(cart__status=4, product__isnull=False)
        .values('product_id', 'product__name', 'product__price', 'product__category__name', 'product__count')
        .annotate(total_units=Sum('count'))
        .order_by('-total_units')[:10]
    )

    # 4. Inventory Value
    all_products = models.Product.objects.all()
    inventory_value = sum(float(p.price) * p.count for p in all_products)
    total_stock_count = sum(p.count for p in all_products)

    context = {
        'monthly_report': monthly_report,
        'categories_perf': categories_perf,
        'top_products_perf': top_products_perf,
        'inventory_value': inventory_value,
        'total_stock_count': total_stock_count,
        'total_products_count': all_products.count(),
    }
    return render(request, 'dashboard/reports.html', context)


@staff_required
def export_orders(request):
    orders_qs = (
        models.Cart.objects.exclude(status=1)
        .select_related('user')
        .prefetch_related('cartproduct_set__product')
        .order_by('-date', 'id')
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar Hisoboti"

    # Styling
    header_fill = PatternFill(start_color="3B5998", end_color="3B5998", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    headers = [
        "№", "Buyurtma ID", "Mijoz Ismi", "Foydalanuvchi nomi",
        "Telefon", "Status", "Moliyaviy Holat", "Oldindan To'lov (%)", "Sana", "Mahsulotlar soni",
        "Jami Summa (so'm)", "To'langan Summa", "Qoldiq Summa", "Manzil"
    ]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_map = {
        1: "Faol savat",
        2: "Yangi (Qabul qilindi)",
        3: "Yo'lda / Jarayonda",
        4: "Yetkazilgan (Tugatildi)",
        5: "Bekor qilingan"
    }

    for idx, order in enumerate(orders_qs, 1):
        order_date = order.date
        if order_date and timezone.is_aware(order_date):
            order_date = timezone.make_naive(order_date).strftime('%d.%m.%Y %H:%M')

        cart_products = order.cartproduct_set.all()
        total_price = float(order.grand_total)
        paid_price = float(order.paid_amount)
        remaining_price = float(order.remaining_amount)
        count_product = sum(cp.count for cp in cart_products)
        user_full_name = order.user.get_full_name() or "—" if order.user else "—"
        user_phone = order.user.phone if hasattr(order.user, 'phone') and order.user.phone else '—'
        user_address = order.user.address if hasattr(order.user, 'address') and order.user.address else '—'

        row = [
            idx,
            order.code,
            user_full_name,
            order.user.username if order.user else "Anonim",
            user_phone,
            status_map.get(order.status, str(order.status)),
            order.get_financial_status_display(),
            f"{order.prepayment_percent}%" if order.prepayment_percent > 0 else "0%",
            order_date,
            count_product,
            total_price,
            paid_price,
            remaining_price,
            user_address
        ]
        ws.append(row)

        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=idx + 1, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            if col_num in (1, 10, 11, 12, 13):
                cell.alignment = Alignment(horizontal="right")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"chimyon_bozor_buyurtmalar_{now().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_required
def export_products_excel(request):
    """
    Barcha mahsulotlar va ombor qoldig'ini Excel (.xlsx) ga eksport qilish.
    """
    products_qs = models.Product.objects.select_related('category').all().order_by('category__name', 'name')
    wb = Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar Katalogi"

    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)

    headers = [
        "№", "Mahsulot Kodi", "Nomi", "Kategoriya",
        "Asosiy Narx (so'm)", "Chegirma Narxi", "Chegirma (%)",
        "Ombordagi Qoldiq (dona)", "Ombor Holati", "Yaratilgan Sana"
    ]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, p in enumerate(products_qs, 1):
        created = p.created_at.strftime('%d.%m.%Y %H:%M') if p.created_at else ''
        stock_status = "Yetarli" if p.count > 5 else ("Kam qolgan" if p.count > 0 else "Tugagan")
        row = [
            idx,
            p.code,
            p.name,
            p.category.name if p.category else '—',
            float(p.price),
            float(p.discount_price) if p.discount_price else '—',
            f"{p.discount_percent}%" if p.discount_percent > 0 else '0%',
            p.count,
            stock_status,
            created
        ]
        ws.append(row)
        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=idx + 1, column=col_num)
            cell.font = data_font

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"chimyon_bozor_mahsulotlar_{now().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_required
def export_inventory_excel(request):
    """
    Ombor qoldiqlari va qiymati bo'yicha maxsus eksport.
    """
    return export_products_excel(request)


# ==========================================
# 8. SAYT SOZLAMALARI VA AUDIT LOGS
# ==========================================

@staff_required
def site_settings_view(request):
    """
    Sayt sozlamalarini Dashboard orqali to'g'ridan-to'g'ri tahrirlash.
    """
    settings_obj = models.SiteSettings.objects.first()
    if not settings_obj:
        settings_obj = models.SiteSettings.objects.create(site_name="Chimyon-bozor")

    if request.method == 'POST':
        try:
            settings_obj.site_name = request.POST.get('site_name', '').strip() or "Chimyon-bozor"
            settings_obj.tagline = request.POST.get('tagline', '').strip()
            settings_obj.phone = request.POST.get('phone', '').strip()
            settings_obj.email = request.POST.get('email', '').strip()
            settings_obj.address = request.POST.get('address', '').strip()
            settings_obj.hero_title = request.POST.get('hero_title', '').strip()
            settings_obj.hero_description = request.POST.get('hero_description', '').strip()
            settings_obj.footer_description = request.POST.get('footer_description', '').strip()
            settings_obj.copyright_text = request.POST.get('copyright_text', '').strip()
            settings_obj.telegram = request.POST.get('telegram', '').strip()
            settings_obj.instagram = request.POST.get('instagram', '').strip()
            settings_obj.facebook = request.POST.get('facebook', '').strip()

            # Prepayment settings
            settings_obj.prepayment_enabled = 'prepayment_enabled' in request.POST
            settings_obj.prepayment_percent = int(request.POST.get('prepayment_percent', 30))
            settings_obj.allowed_prepayment_percentages = request.POST.get('allowed_prepayment_percentages', '30,50,100').strip() or '30,50,100'
            settings_obj.allow_cash_balance = 'allow_cash_balance' in request.POST
            settings_obj.allow_online_balance_payment = 'allow_online_balance_payment' in request.POST

            logo = request.FILES.get('logo')
            if logo:
                validate_image_file(logo)
                settings_obj.logo = logo

            favicon = request.FILES.get('favicon')
            if favicon:
                validate_image_file(favicon)
                settings_obj.favicon = favicon

            settings_obj.save()
            log_admin_action(request, "SETTINGS_UPDATE", "Sayt sozlamalari yangilandi")
            messages.success(request, "Sayt sozlamalari muvaffaqiyatli saqlandi!")
            return redirect('d_settings')
        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {e}")

    return render(request, 'dashboard/settings.html', {'settings': settings_obj})


@staff_required
def audit_logs_view(request):
    """
    Admin harakatlari jurnali (Audit Log).
    """
    query = request.GET.get('query', '').strip()
    action_filter = request.GET.get('action')

    logs = models.AuditLog.objects.select_related('user').all().order_by('-created_at')

    if query:
        logs = logs.filter(
            Q(details__icontains=query) |
            Q(user__username__icontains=query) |
            Q(ip_address__icontains=query)
        )

    if action_filter:
        logs = logs.filter(action=action_filter)

    paginator = Paginator(logs, 25)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    action_types = models.AuditLog.objects.values_list('action', flat=True).distinct()

    return render(request, 'dashboard/audit_logs.html', {
        'logs': page_obj,
        'page_obj': page_obj,
        'query': query,
        'action_filter': action_filter,
        'action_types': action_types,
    })


# ==========================================
# 9. AUTHENTICATION & CHART API
# ==========================================

def log_in(request):
    if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
        return redirect('d_index')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        user = authenticate(username=username, password=password)
        if user is not None and (user.is_staff or user.is_superuser):
            login(request, user)
            log_admin_action(request, "ADMIN_LOGIN", f"Admin tizimga kirdi: {user.username}")
            messages.success(request, f"Xush kelibsiz, {user.username}!")
            next_url = request.GET.get('next') or 'd_index'
            return redirect(next_url)
        messages.error(request, "Foydalanuvchi nomi yoki parol noto'g'ri yoxud admin ruxsati yo'q.")
        return redirect('d_login')

    return render(request, 'dashboard/login.html')


@require_POST
def log_out(request):
    if request.user.is_authenticated:
        log_admin_action(request, "ADMIN_LOGOUT", f"Admin tizimdan chiqdi: {request.user.username}")
    logout(request)
    messages.info(request, "Tizimdan muvaffaqiyatli chiqdingiz.")
    return redirect('d_login')


@staff_required
def revenue_chart_data(request):
    period = request.GET.get('period', '12months')
    current_time = now()

    delivered_sales = models.CartProduct.objects.filter(
        cart__status=4,
        product__isnull=False
    ).select_related('product', 'cart')

    if period == '7days':
        start_date = current_time - timedelta(days=7)
        delivered_sales = delivered_sales.filter(cart__date__gte=start_date)
        daily = defaultdict(float)
        for cp in delivered_sales:
            if cp.cart and cp.cart.date:
                d_str = cp.cart.date.strftime('%d.%m')
                daily[d_str] += float(cp.total_price)
        labels = list(daily.keys())
        totals = list(daily.values())
    elif period == '30days':
        start_date = current_time - timedelta(days=30)
        delivered_sales = delivered_sales.filter(cart__date__gte=start_date)
        daily = defaultdict(float)
        for cp in delivered_sales:
            if cp.cart and cp.cart.date:
                d_str = cp.cart.date.strftime('%d.%m')
                daily[d_str] += float(cp.total_price)
        labels = list(daily.keys())
        totals = list(daily.values())
    else:  # 12months default
        monthly = defaultdict(float)
        for item in delivered_sales:
            if item.cart and item.cart.date:
                month = item.cart.date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                monthly[month] += float(item.total_price)

        ordered_months = sorted(monthly.keys())
        month_names_uz = {
            1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun",
            7: "Iyul", 8: "Avgust", 9: "Sentyabr", 10: "Oktyabr", 11: "Noyabr", 12: "Dekabr"
        }
        labels = [month_names_uz.get(m.month, m.strftime('%B')) for m in ordered_months]
        totals = [monthly[m] for m in ordered_months]

    stats = {
        'total_customers': models.User.objects.count(),
        'total_income': sum(totals),
        'completed_orders': models.Cart.objects.filter(status=4).count(),
        'new_customers': models.User.objects.filter(
            date_joined__month=now().month
        ).count(),
    }

    return JsonResponse({'labels': labels, 'totals': totals, 'stats': stats})

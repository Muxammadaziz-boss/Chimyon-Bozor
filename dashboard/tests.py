from io import BytesIO
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from main.models import Cart, CartProduct, Category, Product, User, OrderStatusHistory, AuditLog, SiteSettings


class DashboardComprehensiveTestCase(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(username='superadmin', password='password123')
        self.staff_user = User.objects.create_user(username='staffmember', password='password123', is_staff=True)
        self.customer = User.objects.create_user(username='customer1', password='password123', is_staff=False)
        
        self.test_image = SimpleUploadedFile(
            name='test_img.jpg',
            content=b'\x47\x49\x46\x38\x39\x61\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00\x21\xf9\x04\x01\x00\x00\x00\x00\x2c\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02\x44\x01\x00\x3b',
            content_type='image/jpeg'
        )

        self.category = Category.objects.create(name='Meva-sabzavotlar', logo='test_logo.png', is_active=True)
        self.product = Product.objects.create(
            category=self.category,
            image='test_prod.png',
            name='Chimyon Olmasi',
            description='Shirin va mazali olma',
            price=10000,
            discount_price=8000,
            discount_status=True,
            count=50,
        )

    # 1. AUTHENTICATION & ACCESS CONTROL
    def test_unauthenticated_user_redirected_to_login(self):
        response = self.client.get(reverse('d_index'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('d_login'), response.url)

    def test_normal_customer_cannot_access_dashboard(self):
        self.client.force_login(self.customer)
        response = self.client.get(reverse('d_index'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('d_login'), response.url)

    def test_staff_user_can_access_dashboard(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(reverse('d_index'))
        self.assertEqual(response.status_code, 200)

    def test_admin_login_and_logout_flow(self):
        response = self.client.post(reverse('d_login'), {
            'username': 'superadmin',
            'password': 'password123'
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('d_index'))

        logout_res = self.client.post(reverse('d_logout'))
        self.assertEqual(logout_res.status_code, 302)
        self.assertRedirects(logout_res, reverse('d_login'))

    # 2. DASHBOARD REAL STATS & ANALYTICS
    def test_dashboard_stats_calculation(self):
        Cart.objects.create(user=self.customer, status=1)  # Active cart - not an order
        c2 = Cart.objects.create(user=self.customer, status=2)  # New
        c3 = Cart.objects.create(user=self.customer, status=3)  # Processing
        c4 = Cart.objects.create(user=self.customer, status=4)  # Completed
        CartProduct.objects.create(cart=c4, product=self.product, count=2)  # 2 * 8000 = 16000

        self.client.force_login(self.admin)
        response = self.client.get(reverse('d_index'))
        self.assertEqual(response.status_code, 200)
        stats = response.context['stats']

        self.assertEqual(stats['total_products'], Product.objects.count())
        self.assertEqual(stats['total_categories'], Category.objects.count())
        self.assertEqual(stats['total_orders'], 3)
        self.assertEqual(stats['new_orders'], 1)
        self.assertEqual(stats['processing_orders'], 1)
        self.assertEqual(stats['completed_orders'], 1)
        self.assertEqual(stats['total_income'], 16000.0)

    def test_sales_analytics_view(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse('d_analytics') + '?period=30days')
        self.assertEqual(response.status_code, 200)
        self.assertIn('period_revenue', response.context)
        self.assertIn('category_sales', response.context)
        self.assertIn('top_products', response.context)

    # 3. CATEGORY CRUD
    def test_category_crud_operations(self):
        self.client.force_login(self.admin)

        # Create
        create_res = self.client.post(reverse('d_create_category'), {
            'name': 'Yangi Kategoriya',
            'logo': self.test_image,
            'is_active': 'on'
        })
        self.assertEqual(create_res.status_code, 302)
        created_cat = Category.objects.get(name='Yangi Kategoriya')
        self.assertTrue(created_cat.is_active)

        # List
        list_res = self.client.get(reverse('d_list_category'))
        self.assertEqual(list_res.status_code, 200)

        # Edit
        edit_res = self.client.post(reverse('d_edit_category', args=[created_cat.id]), {
            'name': 'Kategoriya Yangilandi',
        })
        self.assertEqual(edit_res.status_code, 302)
        created_cat.refresh_from_db()
        self.assertEqual(created_cat.name, 'Kategoriya Yangilandi')
        self.assertFalse(created_cat.is_active)

        # Delete
        del_res = self.client.get(reverse('d_delete_category', args=[created_cat.id]))
        self.assertEqual(del_res.status_code, 302)
        self.assertFalse(Category.objects.filter(id=created_cat.id).exists())

    # 4. PRODUCT CRUD & INVENTORY
    def test_product_crud_operations(self):
        self.client.force_login(self.admin)

        # Create
        create_res = self.client.post(reverse('d_create_product'), {
            'name': 'Yangi Mahsulot',
            'category': self.category.id,
            'description': 'Tavsif',
            'price': '25000',
            'discount_price': '20000',
            'discount_status': 'on',
            'count': '15',
            'image': self.test_image
        })
        self.assertEqual(create_res.status_code, 302)
        created_prod = Product.objects.get(name='Yangi Mahsulot')
        self.assertEqual(created_prod.active_price, 20000)

        # Edit
        edit_res = self.client.post(reverse('d_edit_product', args=[created_prod.code]), {
            'name': 'Mahsulot Yangilandi',
            'category': self.category.id,
            'description': 'Yangi tavsif',
            'price': '30000',
            'count': '20'
        })
        self.assertEqual(edit_res.status_code, 302)
        created_prod.refresh_from_db()
        self.assertEqual(created_prod.name, 'Mahsulot Yangilandi')
        self.assertEqual(created_prod.price, 30000)

        # Delete
        del_res = self.client.get(reverse('d_delete_product', args=[created_prod.code]))
        self.assertEqual(del_res.status_code, 302)
        self.assertFalse(Product.objects.filter(code=created_prod.code).exists())

    def test_inventory_list_and_quick_update(self):
        self.client.force_login(self.admin)
        inv_res = self.client.get(reverse('d_inventory'))
        self.assertEqual(inv_res.status_code, 200)
        self.assertIn('total_stock_units', inv_res.context)

        # Quick stock update
        update_res = self.client.post(reverse('d_update_stock', args=[self.product.code]), {
            'count': '75'
        })
        self.assertEqual(update_res.status_code, 302)
        self.product.refresh_from_db()
        self.assertEqual(self.product.count, 75)

    # 5. ORDER MANAGEMENT, STATUS HISTORY & STOCK REPLENISHMENT
    def test_order_management_workflow_and_stock_return(self):
        self.client.force_login(self.admin)
        initial_stock = self.product.count
        order = Cart.objects.create(user=self.customer, status=2)
        CartProduct.objects.create(cart=order, product=self.product, count=5)

        # Order list with period filter
        list_res = self.client.get(reverse('d_orders') + '?period=today')
        self.assertEqual(list_res.status_code, 200)
        self.assertIn('available_months', list_res.context)

        # Order detail
        detail_res = self.client.get(reverse('d_detail_orders', args=[order.code]))
        self.assertEqual(detail_res.status_code, 200)
        self.assertEqual(detail_res.context['total_amount'], 40000)

        # Update status step (2 -> 3) with comment
        update_res = self.client.post(reverse('d_update_status', args=[order.code]), {
            'target_status': '3',
            'comment': 'Kuryerga topshirildi'
        })
        self.assertEqual(update_res.status_code, 302)
        order.refresh_from_db()
        self.assertEqual(order.status, 3)
        self.assertTrue(OrderStatusHistory.objects.filter(order=order, new_status=3).exists())

        # Cancel order (3 -> 5) and verify stock is replenished
        cancel_res = self.client.post(reverse('d_update_status', args=[order.code]), {
            'target_status': '5',
            'comment': 'Mijoz talabi bilan bekor qilindi'
        })
        self.assertEqual(cancel_res.status_code, 302)
        order.refresh_from_db()
        self.assertEqual(order.status, 5)
        self.product.refresh_from_db()
        self.assertEqual(self.product.count, initial_stock + 5)

    # 6. USER MANAGEMENT & CUSTOMER DOSSIER
    def test_user_management_and_customer_detail(self):
        self.client.force_login(self.admin)
        list_res = self.client.get(reverse('d_list_users'))
        self.assertEqual(list_res.status_code, 200)

        # Customer detail
        detail_res = self.client.get(reverse('d_customer_detail', args=[self.customer.id]))
        self.assertEqual(detail_res.status_code, 200)
        self.assertEqual(detail_res.context['customer'], self.customer)

        # Toggle customer status (active -> inactive)
        toggle_res = self.client.get(reverse('d_toggle_user', args=[self.customer.id]))
        self.assertEqual(toggle_res.status_code, 302)
        self.customer.refresh_from_db()
        self.assertFalse(self.customer.is_active)

        # Cannot toggle self
        self_toggle_res = self.client.get(reverse('d_toggle_user', args=[self.admin.id]))
        self.assertEqual(self_toggle_res.status_code, 302)
        self.admin.refresh_from_db()
        self.assertTrue(self.admin.is_active)

    # 7. REPORTS & EXPORT EXCEL
    def test_reports_and_export_excel(self):
        order = Cart.objects.create(user=self.customer, status=4)
        CartProduct.objects.create(cart=order, product=self.product, count=1)

        self.client.force_login(self.admin)

        # Reports page
        reports_res = self.client.get(reverse('d_reports'))
        self.assertEqual(reports_res.status_code, 200)

        # Export Orders Excel
        response = self.client.get(reverse('d_export_orders'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertGreaterEqual(sheet.max_row, 2)

        # Export Products Excel
        prod_exp = self.client.get(reverse('d_export_products'))
        self.assertEqual(prod_exp.status_code, 200)

    # 8. SETTINGS & AUDIT LOGS
    def test_site_settings_and_audit_logs(self):
        self.client.force_login(self.admin)

        # View settings
        settings_res = self.client.get(reverse('d_settings'))
        self.assertEqual(settings_res.status_code, 200)

        # Update settings
        update_res = self.client.post(reverse('d_settings'), {
            'site_name': 'Chimyon-bozor Yangi',
            'phone': '+998901112233',
            'email': 'admin@chimyon.uz'
        })
        self.assertEqual(update_res.status_code, 302)
        site_st = SiteSettings.objects.first()
        self.assertEqual(site_st.site_name, 'Chimyon-bozor Yangi')

        # Audit logs view
        audit_res = self.client.get(reverse('d_audit_logs'))
        self.assertEqual(audit_res.status_code, 200)
        self.assertTrue(AuditLog.objects.filter(action='SETTINGS_UPDATE').exists())
